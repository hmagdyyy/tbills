import io
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from datetime import datetime

st.set_page_config(page_title="T-Bills NAV Helper", layout="wide")
st.title("T-Bills NAV Helper (Multi-Fund Workbook)")


# -----------------------------------------
# Read fund name from B4 for a sheet
# -----------------------------------------
def get_fund_name(xls: pd.ExcelFile, sheet_name: str) -> str:
    raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    if raw.shape[0] > 3 and raw.shape[1] > 1:
        val = raw.iloc[3, 1]  # B4 (0-based: row 3, col 1)
        if pd.notna(val):
            return str(val).strip()
    return sheet_name


# -----------------------------------------
# Process a single sheet into a DataFrame (for preview)
# -----------------------------------------
def process_sheet(xls: pd.ExcelFile, sheet_name: str, nav: float) -> pd.DataFrame:
    # --- Get TodayDate from B3 (row 3, col B: index [2,1]) ---
    raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    today_raw = raw.iloc[2, 1] if raw.shape[0] > 2 and raw.shape[1] > 1 else None
    today_date = pd.to_datetime(today_raw, errors="coerce", dayfirst=True)

    # --- Main table starting row 8 (header row index = 7) ---
    df = pd.read_excel(xls, sheet_name=sheet_name, header=7)

    # Keep only rows with a Bank
    df = df.dropna(subset=["Bank"])

    # Numeric cols in original sheet
    numeric_cols = [
        "FaceValue", "RatePercent", "NumOfDays",
        "RemainPeriod", "PaidValue", "CurrentAccrude", "Tax"
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Make RatePercent a fraction (e.g. 27 -> 0.27) but keep full precision
    if "RatePercent" in df.columns:
        df["RatePercent"] = df["RatePercent"] / 100.0

    # Date columns – enforce dd/mm/yyyy (day-first)
    for c in ["PurchaseDate", "MaturityDate"]:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce", dayfirst=True)

    # Rename PaidValue -> CurrentValue
    df.rename(columns={"PaidValue": "CurrentValue"}, inplace=True)

    # NAV
    df["NAV"] = nav

    # After Tax = RatePercent * 0.8
    df["After Tax"] = df["RatePercent"] * 0.8

    # Price = 365 / ((NumOfDays * RatePercent) + 365)
    df["Price"] = 365 / ((df["NumOfDays"] * df["RatePercent"]) + 365)

    # PX = Price rounded to 5dp (preview)
    df["PX"] = df["Price"].round(5)

    # ✅ PV = FaceValue * PX  (NOT from sheet)
    df["PV"] = df["FaceValue"] * df["PX"]

    # WeightFromNAV = PV / NAV
    df["WeightFromNAV"] = df["PV"] / nav

    # AvgReturn = After Tax * WeightFromNAV
    df["AvgReturn"] = df["After Tax"] * df["WeightFromNAV"]

    # TodayDate
    df["TodayDate"] = today_date

    # Avg # Of Days = WeightFromNAV * RemainPeriod
    df["Avg # Of Days"] = df["WeightFromNAV"] * df["RemainPeriod"]

    # Accruals = (100 - (Price*100)) * (TodayDate - PurchaseDate) / NumOfDays
    days_diff = (df["TodayDate"] - df["PurchaseDate"]).dt.days
    df["Accruals"] = (100 - (df["Price"] * 100)) * (days_diff / df["NumOfDays"])

    # New PX = (PX*100) + Accruals
    df["New PX"] = (df["PX"] * 100) + df["Accruals"]

    # Breakeven = 365 * ((100 / New PX) - 1) / RemainPeriod
    df["Breakeven"] = 365 * ((100 / df["New PX"]) - 1) / df["RemainPeriod"]

    # Sort by ascending remaining period (ignore TOTAL row for now)
    if "RemainPeriod" in df.columns:
        df = df.sort_values("RemainPeriod", ascending=True)

    # Rounding for preview (Excel keeps formulas accurate)
    num_cols_preview = df.select_dtypes(include=["float64", "int64"]).columns
    for col in num_cols_preview:
        if col == "PX":
            df[col] = df[col].round(5)
        elif col == "Breakeven":
            df[col] = df[col].round(3)
        elif col == "RatePercent":
            # keep full precision
            continue
        else:
            df[col] = df[col].round(2)

    # TOTAL row (preview only)
    sum_row = {
        "Bank": "TOTAL",
        "FaceValue": df["FaceValue"].sum() if "FaceValue" in df.columns else None,
        "PV": df["PV"].sum() if "PV" in df.columns else None,
        "AvgReturn": df["AvgReturn"].sum() if "AvgReturn" in df.columns else None,
        "Avg # Of Days": df["Avg # Of Days"].sum() if "Avg # Of Days" in df.columns else None,
        "CurrentValue": df["CurrentValue"].sum() if "CurrentValue" in df.columns else None,
        "CurrentAccrude": df["CurrentAccrude"].sum() if "CurrentAccrude" in df.columns else None,
        "Tax": df["Tax"].sum() if "Tax" in df.columns else None,
        "NAV": nav,
    }
    df = pd.concat([df, pd.DataFrame([sum_row])], ignore_index=True)

    # Final column order
    final_cols = [
        "Bank",
        "NAV",
        "FaceValue",
        "PV",              # ✅ calculated from FaceValue * PX
        "RatePercent",     # % format in Excel
        "After Tax",       # % format in Excel
        "PurchaseDate",
        "MaturityDate",
        "NumOfDays",
        "Price",
        "PX",
        "TodayDate",
        "RemainPeriod",
        "Breakeven",       # % format in Excel, 3dp
        "CurrentValue",
        "New PX",
        "Accruals",
        "CurrentAccrude",
        "WeightFromNAV",   # % format in Excel
        "AvgReturn",
        "Avg # Of Days",
        "Tax",
    ]
    final_cols = [c for c in final_cols if c in df.columns]
    return df[final_cols]


# -----------------------------------------
# Build Excel with formulas & formatting
# -----------------------------------------
def to_excel_bytes(df: pd.DataFrame, nav: float) -> bytes:
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Result"

    final_cols = list(df.columns)

    # Header
    for col_idx, col_name in enumerate(final_cols, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    data_df = df[df["Bank"] != "TOTAL"].reset_index(drop=True)
    rows_no_total = len(data_df)
    last_data_row = rows_no_total + 1    # rows 2..last_data_row
    total_row_idx = last_data_row + 1    # TOTAL row index

    col_index = {name: i + 1 for i, name in enumerate(final_cols)}

    for i, row in data_df.iterrows():
        r = i + 2

        # Basic values
        ws.cell(row=r, column=col_index["Bank"], value=row["Bank"])

        for name in ["FaceValue", "RatePercent", "NumOfDays", "RemainPeriod",
                     "CurrentValue", "CurrentAccrude", "Tax"]:
            if name in col_index:
                ws.cell(row=r, column=col_index[name], value=row.get(name))

        # Dates (store as true dates)
        for cname in ["PurchaseDate", "TodayDate", "MaturityDate"]:
            if cname in col_index:
                val = row.get(cname)
                ws.cell(row=r, column=col_index[cname], value=val if pd.notna(val) else None)

        # NAV column (same NAV for all rows)
        nav_col = col_index["NAV"]
        if r == 2:
            ws.cell(row=r, column=nav_col, value=nav)
        else:
            ws.cell(row=r, column=nav_col, value=f"={ws.cell(row=2, column=nav_col).coordinate}")

        # After Tax = RatePercent * 0.8
        if "After Tax" in col_index:
            c_rate = ws.cell(row=r, column=col_index["RatePercent"]).coordinate
            ws.cell(row=r, column=col_index["After Tax"], value=f"={c_rate}*0.8")

        # Price = 365 / ((NumOfDays * RatePercent) + 365)
        if "Price" in col_index:
            c_days = ws.cell(row=r, column=col_index["NumOfDays"]).coordinate
            c_rate = ws.cell(row=r, column=col_index["RatePercent"]).coordinate
            ws.cell(row=r, column=col_index["Price"], value=f"=365/(({c_days}*{c_rate})+365)")

        # PX = ROUND(Price, 5)
        if "PX" in col_index and "Price" in col_index:
            c_price = ws.cell(row=r, column=col_index["Price"]).coordinate
            ws.cell(row=r, column=col_index["PX"], value=f"=ROUND({c_price},5)")

        # ✅ PV = FaceValue * PX (formula)
        if "PV" in col_index:
            c_face = ws.cell(row=r, column=col_index["FaceValue"]).coordinate
            c_px = ws.cell(row=r, column=col_index["PX"]).coordinate
            ws.cell(row=r, column=col_index["PV"], value=f"={c_face}*{c_px}")

        # WeightFromNAV = PV / NAV
        if "WeightFromNAV" in col_index:
            c_pv = ws.cell(row=r, column=col_index["PV"]).coordinate
            c_nav = ws.cell(row=r, column=nav_col).coordinate
            ws.cell(row=r, column=col_index["WeightFromNAV"], value=f"={c_pv}/{c_nav}")

        # AvgReturn = After Tax * WeightFromNAV
        if "AvgReturn" in col_index:
            c_at = ws.cell(row=r, column=col_index["After Tax"]).coordinate
            c_w = ws.cell(row=r, column=col_index["WeightFromNAV"]).coordinate
            ws.cell(row=r, column=col_index["AvgReturn"], value=f"={c_at}*{c_w}")

        # Avg # Of Days = WeightFromNAV * RemainPeriod
        if "Avg # Of Days" in col_index:
            c_w = ws.cell(row=r, column=col_index["WeightFromNAV"]).coordinate
            c_rem = ws.cell(row=r, column=col_index["RemainPeriod"]).coordinate
            ws.cell(row=r, column=col_index["Avg # Of Days"], value=f"={c_w}*{c_rem}")

        # Accruals = (100 - (Price*100)) * ((TodayDate - PurchaseDate)/NumOfDays)
        if "Accruals" in col_index:
            c_price = ws.cell(row=r, column=col_index["Price"]).coordinate
            c_today = ws.cell(row=r, column=col_index["TodayDate"]).coordinate
            c_pur = ws.cell(row=r, column=col_index["PurchaseDate"]).coordinate
            c_days = ws.cell(row=r, column=col_index["NumOfDays"]).coordinate
            ws.cell(row=r, column=col_index["Accruals"], value=f"=(100-({c_price}*100))*(({c_today}-{c_pur})/{c_days})")

        # New PX = (PX*100) + Accruals
        if "New PX" in col_index:
            c_px = ws.cell(row=r, column=col_index["PX"]).coordinate
            c_accr = ws.cell(row=r, column=col_index["Accruals"]).coordinate
            ws.cell(row=r, column=col_index["New PX"], value=f"=({c_px}*100)+{c_accr}")

        # Breakeven = 365 * ((100 / New PX) - 1) / RemainPeriod
        if "Breakeven" in col_index:
            c_newpx = ws.cell(row=r, column=col_index["New PX"]).coordinate
            c_rem = ws.cell(row=r, column=col_index["RemainPeriod"]).coordinate
            ws.cell(row=r, column=col_index["Breakeven"], value=f"=365*((100/{c_newpx})-1)/{c_rem}")

    # TOTAL row with SUM formulas
    ws.cell(row=total_row_idx, column=col_index["Bank"], value="TOTAL")

    def add_sum(col_name: str):
        c_idx = col_index[col_name]
        col_letter = ws.cell(row=1, column=c_idx).column_letter
        ws.cell(row=total_row_idx, column=c_idx, value=f"=SUM({col_letter}2:{col_letter}{last_data_row})")

    for cname in ["FaceValue", "PV", "AvgReturn", "Avg # Of Days",
                  "CurrentValue", "CurrentAccrude", "Tax"]:
        if cname in col_index:
            add_sum(cname)

    # NAV in TOTAL row = NAV in first data row
    ws.cell(row=total_row_idx, column=col_index["NAV"],
            value=f"={ws.cell(row=2, column=col_index['NAV']).coordinate}")

    # ---------------- Formatting ----------------
    number_fmt_2 = '#,##0.00'
    number_fmt_5 = '#,##0.00000'

    # Non-percentage numeric columns
    numeric_columns = [
        "FaceValue", "PV", "Price", "PX", "Accruals", "New PX",
        "NumOfDays", "RemainPeriod", "AvgReturn", "Avg # Of Days",
        "CurrentValue", "CurrentAccrude", "Tax", "NAV"
    ]
    for cname in numeric_columns:
        if cname not in col_index:
            continue
        c_idx = col_index[cname]
        fmt = number_fmt_5 if cname == "PX" else number_fmt_2
        for rr in range(2, total_row_idx + 1):
            cell = ws.cell(row=rr, column=c_idx)
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.number_format = fmt

    # Percent formats
    if "RatePercent" in col_index:
        c_idx = col_index["RatePercent"]
        for rr in range(2, total_row_idx + 1):
            ws.cell(row=rr, column=c_idx).number_format = "0.0000%"

    if "After Tax" in col_index:
        c_idx = col_index["After Tax"]
        for rr in range(2, total_row_idx + 1):
            ws.cell(row=rr, column=c_idx).number_format = "0.00%"

    if "WeightFromNAV" in col_index:
        c_idx = col_index["WeightFromNAV"]
        for rr in range(2, total_row_idx + 1):
            ws.cell(row=rr, column=c_idx).number_format = "0.00%"

    if "Breakeven" in col_index:
        c_idx = col_index["Breakeven"]
        for rr in range(2, total_row_idx + 1):
            ws.cell(row=rr, column=c_idx).number_format = "0.000%"

    # Dates as dd/mm/yyyy
    date_fmt = "DD/MM/YYYY"
    for cname in ["PurchaseDate", "TodayDate", "MaturityDate"]:
        if cname in col_index:
            c_idx = col_index[cname]
            for rr in range(2, total_row_idx + 1):
                cell = ws.cell(row=rr, column=c_idx)
                if isinstance(cell.value, datetime):
                    cell.number_format = date_fmt

    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# -----------------------------------------
# Streamlit UI
# -----------------------------------------
uploaded_file = st.file_uploader("Upload Excel workbook", type=["xls", "xlsx"])

if uploaded_file:
    xls = pd.ExcelFile(uploaded_file)
    sheet_names = xls.sheet_names

    st.markdown(f"**Found {len(sheet_names)} sheet(s):** {', '.join(sheet_names)}")

    for sheet_name in sheet_names:
        fund_name = get_fund_name(xls, sheet_name)

        with st.expander(f"Fund: {fund_name} (sheet: {sheet_name})", expanded=False):
            nav_value = st.number_input(
                f"NAV for {fund_name}",
                min_value=0.01,
                step=100.0,
                format="%.2f",
                key=f"nav_{sheet_name}",
            )

            if nav_value and nav_value > 0:
                df_result = process_sheet(xls, sheet_name, nav_value)

                st.write("Preview:")
                st.dataframe(df_result, use_container_width=True)

                excel_data = to_excel_bytes(df_result, nav_value)

                safe_fund_name = fund_name.replace(" ", "_") if fund_name else sheet_name
                st.download_button(
                    label=f"📥 Download Excel for {fund_name}",
                    data=excel_data,
                    file_name=f"{safe_fund_name}_t_bills_nav.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{sheet_name}",
                )
else:
    st.info("Upload a workbook with one or more T-Bills sheets to begin.")

